import streamlit as st
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
import matplotlib.animation as animation
import io
from openpyxl import load_workbook

# Funções de interpolação (mantidas para o pré-cálculo)
def create_interpolation_functions(ks_data, cs_data, ds_data, hg_data):
    ks_interp = interp1d(ks_data[0], ks_data[1], kind='linear', fill_value="extrapolate")
    cs_interp = interp1d(cs_data[0], cs_data[1], kind='linear', fill_value="extrapolate")
    ds_interp = interp1d(ds_data[0], ds_data[1], kind='linear', fill_value="extrapolate")
    hg_interp = interp1d(hg_data[0], hg_data[1], kind='linear', fill_value="extrapolate")
    return ks_interp, cs_interp, ds_interp, hg_interp

# Função para pré-calcular propriedades interpoladas
def precalculate_interpolated_properties(ks_data, cs_data, ds_data, hg_data, temp_range_start=20, temp_range_end=1200, temp_step=10): #alterei inicio para 20 para evitar problemas com interpolação abaixo de 20
    ks_interp, cs_interp, ds_interp, hg_interp = create_interpolation_functions(ks_data, cs_data, ds_data, hg_data)
    precalc_temps = np.arange(temp_range_start, temp_range_end + temp_step, temp_step)
    precalc_ks = ks_interp(precalc_temps)
    precalc_cs = cs_interp(precalc_temps)
    precalc_ds = ds_interp(precalc_temps)
    precalc_hg = hg_interp(precalc_temps)
    return precalc_temps, precalc_ks, precalc_cs, precalc_ds, precalc_hg

# Função para busca rápida e interpolação linear
def get_interpolated_property(temp, precalc_temps, precalc_values):
    if temp <= precalc_temps[0]:
        return precalc_values[0]
    if temp >= precalc_temps[-1]:
        return precalc_values[-1]

    index = np.searchsorted(precalc_temps, temp, side="right")
    if index > 0 and index < len(precalc_temps):
        temp_baixo = precalc_temps[index-1]
        temp_alto = precalc_temps[index]
        valor_baixo = precalc_values[index-1]
        valor_alto = precalc_values[index]
        # Interpolação linear
        return valor_baixo + (valor_alto - valor_baixo) * (temp - temp_baixo) / (temp_alto - temp_baixo)
    elif index == 0:
        return precalc_values[0]
    else: # index == len(precalc_temps)
        return precalc_values[-1]

# Função para a calculadora de temperatura na cantoneira
def angle_heat_transfer_calculator(min_temp, len_x, len_y, len_w, len_z, delta,
                                        precalc_temps, precalc_ks, precalc_cs, precalc_ds, precalc_hg, # Dados pré-calculados
                                        ks_data, cs_data, ds_data, hg_data, # Mantenha os dados originais para dt 
                                        tamb, t_initial):

    # Dimensões da grade
    nx = int(len_x / delta) + 1
    ny = int(len_y / delta) + 1
    nw = int(len_w / delta) + 1
    nz = int(len_z / delta) + 1

    print(f"nx: {nx}, ny: {ny}, nw: {nw}, nz: {nz}")

    
    domain = np.full((nx, ny), 1, dtype=np.float64)
    for i in range(nx):
        for j in range(ny):
            if (j == 0):
                domain[i, j] = 3
            if (i == 0):
                domain[i, j] = 3
            if (i > (nz-1) and j * delta > len_w):
                domain[i, j] = 0
            if (j == (ny-1) and i * delta < len_z):
                domain[i, j] = 3
            if (i == (nx-1) and j * delta < len_w):
                domain[i, j] = 3
            if (i == (nz-1) and j * delta > len_w):
                domain[i, j] = 3
            if (j == (nw-1) and i * delta > len_z):
                domain[i, j] = 3  

    # Inicialização da matriz de temperatura
    t = np.zeros((nx, ny), dtype=np.float64)
    if isinstance(t_initial, (int, float)):
        t[:] = t_initial
    else:
        t = np.copy(t_initial) 

    t_prev = np.copy(t)

    # Pontos para monitorar a temperatura
    pontos = [(0, int(ny/2)), (int(nz/2), int(2*ny/3)), (int(2*nx/3) , int(nw/2))]
    temperaturas = {p: [] for p in pontos}

    # Lista para armazenar os campos de temperatura para a animação
    temperature_fields = []  # Aqui está a lista para a animação

    tempo = 0  # Inicializa o tempo
    dt = 0.001  # Valor inicial para dt, será atualizado dentro do loop
    dts = []

    # Crie um espaço reservado no Streamlit para exibir as mensagens
    status_text = st.empty()


    # Loop no tempo
    while True:
        # Imprime a temperatura máxima e o passo de tempo
        max_temp_atual = np.min(temperaturas[pontos[1]]) if temperaturas[pontos[1]] else t.max()
        message = f"Time: {tempo:.2f} sec, Leg 1 Middle Temp: {max_temp_atual:.2f} °C"
        status_text.info(message)  # Exibe a mensagem no Streamlit

        # Verifica se a temperatura máxima atingiu o valor desejado
        if max_temp_atual <= min_temp:
            print(f"Temperatura máxima atingida ({max_temp_atual:.2f} °C) no tempo {tempo:.2f} s")
            break
       
       
        # Loop espacial
        for i in range(0, nx):
            for j in range(0, ny):
                if domain[i, j] != 0:  # Calcula apenas dentro do domínio
         
                    ks = get_interpolated_property(t_prev[i, j], precalc_temps, precalc_ks)
                    cs = get_interpolated_property(t_prev[i, j], precalc_temps, precalc_cs)
                    ds = get_interpolated_property(t_prev[i, j], precalc_temps, precalc_ds)
                    hg = get_interpolated_property(t_prev[i, j], precalc_temps, precalc_hg)

                    # Cálculo da difusividade térmica
                    alpha = ks / (cs * ds)

                    # Definição do número de Biot (Ms)
                    ms = hg * (delta / 1000) / ks

                    # Definição do número de Fourier (Ml)
                    ml = max(4, ms + 3, 2 * ms + 2) #PODE SER QUE PRECISE AUMENTAR ISSO AQUI

                    # Cálculo do passo de tempo (agora dentro do loop, pois depende da temperatura)
                    dt = (delta / 1000)**2 / (ml * alpha)
                    if (dt > 1): dt = 1  #Limitando dt para no máximo 1 segundo

                    # Condição interna
                    if i > 0 and i < nx - 1 and j > 0 and j < ny - 1 and domain[i,j] == 1:
                        t[i, j] = (1 / ml) * (t_prev[i + 1, j] + t_prev[i - 1, j] + t_prev[i, j + 1] + t_prev[i, j - 1]) + t_prev[i, j] * (ml - 4) / ml
                    # Condições de contorno (convecção)
                    else:
                        # Superfície externa Flange
                        if i == 0 and j < ny - 1 and domain[i,j] == 3:
                            t[i, j] = (ms / ml) * tamb + (1 / ml) * (t_prev[i, j - 1] + t_prev[i, j + 1] + t_prev[i + 1, j]) + t_prev[i, j] * (1 - (ms + 3) / ml)
                        # superficie espessura leg 2
                        if i == nx - 1 and domain[i, j] == 3:
                            t[i, j] = (ms / ml) * tamb + (1 / ml) * (t_prev[i, j - 1] + t_prev[i, j + 1] + t_prev[i - 1, j]) + t_prev[i, j] * (1 - (ms + 3) / ml)
                        # superficie externa leg 2
                        if j == 0 and i < nx - 1 and domain[i, j] == 3:
                            t[i, j] = (ms / ml) * tamb + (1 / ml) * (t_prev[i - 1, j] + t_prev[i + 1, j] + t_prev[i, j + 1]) + t_prev[i, j] * (1 - (ms + 3) / ml)
                        # Superfície superior Flange
                        if j == ny - 1 and domain[i, j] == 3:
                            t[i, j] = (ms / ml) * tamb + (1 / ml) * (t_prev[i - 1, j] + t_prev[i + 1, j] + t_prev[i, j - 1]) + t_prev[i, j] * (1 - (ms + 3) / ml)
                        # Canto inferior leg 1 e leg 2
                        if i == 0 and j == 0 and domain[i, j] == 3:
                            t[i, j] = 2 * (ms / ml) * tamb + (1 / ml) * (t_prev[i, j + 1] + t_prev[i + 1, j]) + t_prev[i, j] * (1 - (2 * ms + 2) / ml)
                        # Canto superior Flange esquerdo
                        if i == 0 and j == ny - 1 and domain[i, j] == 3:
                            t[i, j] = 2 * (ms / ml) * tamb + (1 / ml) * (t_prev[i, j - 1] + t_prev[i + 1, j]) + t_prev[i, j] * (1 - (2 * ms + 2) / ml)
                         # Canto inferior leg 2
                        if i == nx - 1 and j == 0 and domain[i, j] == 3:
                            t[i, j] = 2 * (ms / ml) * tamb + (1 / ml) * (t_prev[i, j + 1] + t_prev[i - 1, j]) + t_prev[i, j] * (1 - (2 * ms + 2) / ml)
                        # Canto plano de simetria superfície interna
                        if i == nx - 1 and j == int(nw-1) and domain[i, j] == 3:
                            t[i, j] = 2 * (ms / ml) * tamb + (1 / ml) * (t_prev[i, j - 1] + t_prev[i - 1, j]) + t_prev[i, j] * (1 - (2 * ms + 2) / ml)
                        # Canto superficie flange interno
                        if i == int(nz-1) and j == ny-1 and domain[i, j] == 3:
                            t[i, j] = 2 * (ms / ml) * tamb + (1 / ml) * (t_prev[i, j - 1] + t_prev[i - 1, j]) + t_prev[i, j] * (1 - (2 * ms + 2) / ml)
                        #Condições de contorno nas bordas internas do L
                        # Borda interna vertical
                        if i == int(nz-1) and j >= int(nw-1) and j < ny - 1 and domain[i, j] == 3:
                            t[i, j] = (ms / ml) * tamb + (1 / ml) * (t_prev[i, j - 1] + t_prev[i, j + 1] + t_prev[i - 1, j]) + t_prev[i, j] * (1 - (ms + 3) / ml)

                        # Borda interna horizontal
                        if j == int(nw-1) and i >= int(nz-1) and i < nx - 1 and domain[i, j] == 3:
                            t[i, j] = (ms / ml) * tamb + (1 / ml) * (t_prev[i - 1, j] + t_prev[i + 1, j] + t_prev[i, j - 1]) + t_prev[i, j] * (1 - (ms + 3) / ml)

        # Armazenar o dt atual
        dts.append(dt)

        # Armazenar temperaturas nos pontos monitorados
        for p in pontos:
            temperaturas[p].append(t[p])

        # Armazenar o campo de temperatura para a animação
        temperature_fields.append(np.copy(t)) # Salva uma cópia do campo de temperatura

        t_prev = np.copy(t)
        tempo += dt  # Incrementa o tempo

        # Debug: Imprimir a matriz de temperatura em alguns pontos do tempo
        #if tempo % (t_final / 10) == 0:
            #print(f"Temperatura no tempo {tempo}:")
            #print(t)

    return temperature_fields, temperaturas, dts, pontos, nx, ny, nz, nw


def h_beam_heat_transfer_calculator(min_temp, len_x, len_y, len_w, len_z, delta,
                                        precalc_temps, precalc_ks, precalc_cs, precalc_ds, precalc_hg, # Dados pré-calculados
                                        ks_data, cs_data, ds_data, hg_data, # Mantenha os dados originais para dt (opcional, pode otimizar isso também)
                                        tamb, t_initial):

    # Dimensões da grade
    nx = int(len_x / delta) + 1
    ny = int(len_y / delta) + 1
    nw = int(len_w / delta) + 1
    nz = int(len_z / delta) + 1

    print(f"nx: {nx}, ny: {ny}, nw: {nw}, nz: {nz}")

    # Criar a geometria em forma de L
    domain = np.full((nx, ny), 1, dtype=np.float64)
    for i in range(nx):
        for j in range(ny):
       
            if (j == 0):
                domain[i, j] = 2
            if (i == 0):
                domain[i, j] = 3
            if (i > (nz-1) and j * delta > len_w):
                domain[i, j] = 0
            if (j == (ny-1) and i * delta < len_z):
                domain[i, j] = 3
            if (i == (nx-1) and j * delta < len_w):
                domain[i, j] = 2
            if (i == (nz-1) and j * delta > len_w):
                domain[i, j] = 3
            if (j == (nw-1) and i * delta > len_z):
                domain[i, j] = 3  

    # Inicialização da matriz de temperatura
    t = np.zeros((nx, ny), dtype=np.float64)
    if isinstance(t_initial, (int, float)):
        t[:] = t_initial
    else:
        t = np.copy(t_initial) 

    t_prev = np.copy(t)

    # Pontos para monitorar a temperatura
    pontos = [(0, 0), (int(nz/2), int(2*ny/3)), (int(nx-1) , 0)]
    temperaturas = {p: [] for p in pontos}

    # Lista para armazenar os campos de temperatura para a animação
    temperature_fields = []  # Aqui está a lista para a animação

    tempo = 0  # Inicializa o tempo
    dt = 0.001  # Valor inicial para dt, será atualizado dentro do loop
    dts = []

    # Crie um espaço reservado no Streamlit para exibir as mensagens
    status_text = st.empty()


    # Loop no tempo
    while True:
        # Imprime a temperatura máxima e o passo de tempo
        max_temp_atual = np.min(temperaturas[pontos[1]]) if temperaturas[pontos[1]] else t.max()
        message = f"Time: {tempo:.2f} sec, Flange Middle Temp: {max_temp_atual:.2f} °C"
        status_text.info(message)  # Exibe a mensagem no Streamlit

        # Verifica se a temperatura máxima atingiu o valor desejado
        if max_temp_atual <= min_temp:
            print(f"Temperatura máxima atingida ({max_temp_atual:.2f} °C) no tempo {tempo:.2f} s")
            break
       
       
        # Loop espacial
        for i in range(0, nx):
            for j in range(0, ny):
                if domain[i, j] != 0:  # Calcula apenas dentro do domínio
                    # Calcula propriedades dependentes da temperatura
                   
                    ks = get_interpolated_property(t_prev[i, j], precalc_temps, precalc_ks)
                    cs = get_interpolated_property(t_prev[i, j], precalc_temps, precalc_cs)
                    ds = get_interpolated_property(t_prev[i, j], precalc_temps, precalc_ds)
                    hg = get_interpolated_property(t_prev[i, j], precalc_temps, precalc_hg)

                    # Cálculo da difusividade térmica
                    alpha = ks / (cs * ds)

                    # Definição do número de Biot (Ms)
                    ms = hg * (delta / 1000) / ks

                    # Definição do número de Fourier (Ml)
                    ml = max(4, ms + 3, 2 * ms + 2) 

                    # Cálculo do passo de tempo (agora dentro do loop, pois depende da temperatura)
                    dt = (delta / 1000)**2 / (ml * alpha)
                    if (dt > 1): dt = 1  #Limitando dt para no máximo 1 segundo

                    # Condição interna
                    if i > 0 and i < nx - 1 and j > 0 and j < ny - 1 and domain[i,j] == 1:
                        t[i, j] = (1 / ml) * (t_prev[i + 1, j] + t_prev[i - 1, j] + t_prev[i, j + 1] + t_prev[i, j - 1]) + t_prev[i, j] * (ml - 4) / ml
                    # Condições de contorno (convecção)
                    else:
                        # Superfície externa Flange
                        if i == 0 and j < ny - 1 and domain[i,j] == 3:
                            t[i, j] = (ms / ml) * tamb + (1 / ml) * (t_prev[i, j - 1] + t_prev[i, j + 1] + t_prev[i + 1, j]) + t_prev[i, j] * (1 - (ms + 3) / ml)
                        # Plano de simetria Vertical
                        if i == nx - 1 and domain[i, j] == 2:
                            t[i, j] = (1 / ml) * (t_prev[i, j - 1] + t_prev[i, j + 1] + t_prev[i - 1, j]) + t_prev[i, j] * (ml - 3) / ml
                        # Plano de Simetria Horizontal
                        if j == 0 and i < nx - 1 and domain[i, j] == 2:
                            t[i, j] = (1 / ml) * (t_prev[i - 1, j] + t_prev[i + 1, j] + t_prev[i, j + 1]) + t_prev[i, j] * (ml - 3) / ml
                        # Superfície superior Flange
                        if j == ny - 1 and domain[i, j] == 3:
                            t[i, j] = (ms / ml) * tamb + (1 / ml) * (t_prev[i - 1, j] + t_prev[i + 1, j] + t_prev[i, j - 1]) + t_prev[i, j] * (1 - (ms + 3) / ml)
                        # Canto meio superfície Flange
                        if i == 0 and j == 0 and domain[i, j] == 3:
                            t[i, j] = (ms / ml) * tamb + (1 / ml) * (t_prev[i, j + 1] + t_prev[i + 1, j]) + t_prev[i, j] * (1 - ( ms + 2) / ml)
                        # Canto superior Flange esquerdo
                        if i == 0 and j == ny - 1 and domain[i, j] == 3:
                            t[i, j] = 2 * (ms / ml) * tamb + (1 / ml) * (t_prev[i, j - 1] + t_prev[i + 1, j]) + t_prev[i, j] * (1 - (2 * ms + 2) / ml)
                         # Canto meio plano de simetria
                        if i == nx - 1 and j == 0 and domain[i, j] == 2:
                            t[i, j] = (1 / ml) * (t_prev[i, j + 1] + t_prev[i - 1, j]) + t_prev[i, j] * (ml - 2) / ml
                        # Canto plano de simetria superfície interna
                        if i == nx - 1 and j == int(nw-1) and domain[i, j] == 3:
                            t[i, j] = (ms / ml) * tamb + (1 / ml) * (t_prev[i, j - 1] + t_prev[i - 1, j]) + t_prev[i, j] * (1 - ( ms + 2) / ml)
                        # Canto superficie flange interno
                        if i == int(nz-1) and j == ny-1 and domain[i, j] == 3:
                            t[i, j] = 2 * (ms / ml) * tamb + (1 / ml) * (t_prev[i, j - 1] + t_prev[i - 1, j]) + t_prev[i, j] * (1 - (2 * ms + 2) / ml)
                        #Condições de contorno nas bordas internas do L
                        # Borda interna vertical
                        if i == int(nz-1) and j >= int(nw-1) and j < ny - 1 and domain[i, j] == 3:
                            t[i, j] = (ms / ml) * tamb + (1 / ml) * (t_prev[i, j - 1] + t_prev[i, j + 1] + t_prev[i - 1, j]) + t_prev[i, j] * (1 - (ms + 3) / ml)

                        # Borda interna horizontal
                        if j == int(nw-1) and i >= int(nz-1) and i < nx - 1 and domain[i, j] == 3:
                            t[i, j] = (ms / ml) * tamb + (1 / ml) * (t_prev[i - 1, j] + t_prev[i + 1, j] + t_prev[i, j - 1]) + t_prev[i, j] * (1 - (ms + 3) / ml)

        # Armazenar o dt atual
        dts.append(dt)

        # Armazenar temperaturas nos pontos monitorados
        for p in pontos:
            temperaturas[p].append(t[p])

        # Armazenar o campo de temperatura para a animação
        temperature_fields.append(np.copy(t)) # Salva uma cópia do campo de temperatura

        t_prev = np.copy(t)
        tempo += dt  # Incrementa o tempo

      
    return temperature_fields, temperaturas, dts, pontos, nx, ny, nz

@st.cache_data
def create_animation_and_temps(temperature_fields, dts, len_x, len_y, delta, min_temp, t_initial, pontos, nx, ny, nz, animation_interval, section_type):
    """
    Cria a animação e calcula os tempos acumulados, usando st.cache_data para armazenar em cache.
    """
    tempos = np.cumsum(dts)

    # Definir os limites da escala de temperatura
    vmin = min_temp
    vmax = t_initial

    # Criar uma nova lista com os campos de temperatura a cada 'animation_interval' segundos
    sampled_temperature_fields = temperature_fields[::animation_interval]

    # Criar a animação
    fig, ax = plt.subplots()
    im = ax.imshow(sampled_temperature_fields[0].T, cmap='hot', origin='lower', extent=[0, len_x, 0, len_y], animated=True, vmin=vmin, vmax=vmax)
    cbar = plt.colorbar(im, label='Temperature (°C)')

    # Adicionar rótulos e títulos específicos para cada seção
    if section_type == "angle":
        ax.set_xlabel(' Leg 2 Length (mm)')
        ax.set_ylabel('Leg 1 Length (mm)')
        ax.set_title('Temperature Distribution Angle Section')
    elif section_type == "h_beam":
        ax.set_xlabel('Half Web Length (mm)')
        ax.set_ylabel('Half Flange Length (mm)')
        ax.set_title('Temperature Distribution H-Beam Section')

    # Adicionar os pontos e rótulos à animação
    point_colors = ['black', 'black', 'black']  # Cores para cada ponto
    
    # Rótulos para os pontos, específicos para cada seção
    if section_type == "angle":
        point_labels_animation = ['Leg 1 Surface', 'Leg 1 Middle', 'Leg 2 Middle']  # Rótulos para os pontos
    elif section_type == "h_beam":
        point_labels_animation = ['Flange Surface', 'Flange Middle', 'Web Middle']
    

    for i, p in enumerate(pontos):
        x_pixel = p[0] * delta
        y_pixel = p[1] * delta
        ax.plot(x_pixel, y_pixel, marker='o', color=point_colors[i], markersize=5)  # Ajuste o tamanho dos marcadores conforme necessário
        # Adicionar o rótulo ao lado dos 3 pontos
        ax.text(x_pixel + 2, y_pixel + 1, point_labels_animation[i], color='black', fontsize=8)  # Ajuste a posição e o tamanho da fonte conforme necessário

    time_text = ax.text(0.75, 0.90, '', transform=ax.transAxes, color='black')  # Posiciona o texto do tempo

    def update_fig(num):
        im.set_array(sampled_temperature_fields[num].T)
        time_text.set_text(f'Time: {tempos[num * animation_interval]:.1f} s')  # Atualiza o texto do tempo
        return [im, time_text]

    ani = animation.FuncAnimation(fig, update_fig, frames=len(sampled_temperature_fields), blit=True)

    # Salvar a animação como um arquivo GIF
    try:
        ani.save('temperature_animation.gif', writer='pillow')  
        print("Animation saved as temperature_animation.gif")
    except Exception as e:
        print(f"Error saving animation (Pillow might not be installed): {e}")

    plt.close(fig) 

    return tempos, "temperature_animation.gif"  # Retorna os tempos e o caminho da animação

# Funções auxiliares
def calculate_cooling_rate(temperatures, times, start_temp, end_temp):
    """Calcula a taxa de resfriamento entre duas temperaturas."""
    try:
        start_index = next(i for i, temp in enumerate(temperatures) if temp <= start_temp)
        end_index = next(i for i, temp in enumerate(temperatures[start_index:]) if temp <= end_temp) + start_index
        time_diff = times[end_index] - times[start_index]
        temp_diff = start_temp - end_temp
        cooling_rate = temp_diff / time_diff if time_diff > 0 else float('inf')
        return cooling_rate
    except StopIteration:
        return float('inf')  # Resfriamento não atingiu a temperatura final

def save_temperature_data(tempos, temperaturas, pontos, temp_interval, target_point_index):
    """
    Salva os dados de temperatura em um arquivo TXT, amostrando os dados a cada 'temp_interval' graus Celsius de variação
    apenas para o ponto especificado por 'target_point_index'.
    """
    output = io.StringIO()
    output.write("Time (s)")
    for p in pontos:
        output.write(f", Point {p}")
    output.write("\n")

    # Converter listas para arrays numpy para facilitar a indexação
    tempos = np.array(tempos)
    temperaturas = {p: np.array(temperaturas[p]) for p in pontos}

    # Identificar o ponto de interesse
    target_point = pontos[target_point_index]

    # Rastrear a última temperatura salva e o índice correspondente PARA O PONTO DE INTERESSE
    last_saved_temp, last_saved_index = temperaturas[target_point][0], 0

    # Sempre salvar o primeiro ponto
    output.write(f"{tempos[0]:.1f}")
    for p in pontos:
        output.write(f", {temperaturas[p][0]:.2f}")
    output.write("\n")

    # Iterar pelos tempos, verificando a variação da temperatura NO PONTO DE INTERESSE
    for i in range(1, len(tempos)):
        temp_change = abs(temperaturas[target_point][i] - last_saved_temp)
        if temp_change >= temp_interval:
            # Salvar o ponto atual
            output.write(f"{tempos[i]:.1f}")
            for p_inner in pontos:
                output.write(f", {temperaturas[p_inner][i]:.2f}")
            output.write("\n")
            # Atualizar a última temperatura salva e o índice
            last_saved_temp, last_saved_index = temperaturas[target_point][i], i

    return output.getvalue()

    
# Configuração da página Streamlit
st.set_page_config(
    page_title="Angle Heat Transfer Calculator",
    #page_icon=":desktop_computer:",
    page_icon=":computer:",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Variáveis de estado da sessão
if 'authentication_status' not in st.session_state:
    st.session_state['authentication_status'] = None
if 'username' not in st.session_state:
    st.session_state['username'] = None

USUARIOS = {
    "roberto.tiburcio": "rtf2679",
    "vinicius.ottani": "vco4285",
    "marcelo.rebellato": "mar4928",
    "antonio.gorni": "aag1958",
    "jose.bacalhau": "jbb1985",
}

def autenticar(username, password):
    """
    Verifica se o usuário e senha fornecidos correspondem às credenciais no dicionário.
    Retorna True se a autenticação for bem-sucedida, False caso contrário.
    """
    if username in USUARIOS and USUARIOS[username] == password:
        return True
    return False

# Página de Login
def login():
    st.title("Login")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")

    if st.button("Sign In"):
        if autenticar(username, password):
            st.session_state['authentication_status'] = True
            st.session_state['username'] = username
            st.success("Login successful! Please choose a section calculator from the menu.")
        else:
            st.session_state['authentication_status'] = False
            st.error("Login failed. Please try again.")

# Página da Calculadora de Seção Angular
def angle_calculator():
    # Limpar as variáveis de sessão específicas do H-Beam
    for key in ['hbeam_tempos', 'hbeam_temperaturas', 'hbeam_pontos', 'hbeam_start_temp', 'hbeam_end_temp', 'hbeam_nx', 'hbeam_ny', 'hbeam_nz', 'hbeam_nw']:
        if key in st.session_state:
            del st.session_state[key]

    st.title("Angle Heat Transfer Calculator")

    st.image("angle_dimensions_complete2.png", use_container_width=True)

    # Opção de entrada de dados
    data_option = st.radio("Select data input type:", ("Typing", "Upload XLSX File"))

    # Inicializar dicionários para os dados
    input_data = {}

    if data_option == "Typing":
        st.subheader("Input Data (Typing)")
        # Dados de entrada numéricos
        input_data['len_y'] = st.number_input("Leg 1 Length  (mm)", value=152.0, key="len_y")
        input_data['len_z'] = st.number_input("Leg 1 Thickness (mm)", value=11.1, key="len_z")
        input_data['len_x'] = st.number_input("Leg 2 Length  (mm)", value=203.0, key="len_x")
        input_data['len_w'] = st.number_input("Leg 2 Thickness (mm)", value=11.1, key="len_w")
        #input_data['delta'] = st.number_input("Mesh Size (mm)", value=5.0, key="delta")
        input_data['min_temp'] = st.number_input("Final Temperature in Leg 1 Middle (°C)", value=750.0, key="min_temp")
        input_data['tamb'] = st.number_input("Environment Temperature (°C)", value=30.0, key="tamb")
        input_data['t_initial'] = st.number_input("Initial Temperature (Finishing Rolling Temperature) (°C)", value=850.0, key="t_initial")

        # Dados de propriedades variáveis (listas de números)
        st.write("Thermal Conductivity (ks_data):")
        input_data['ks_temperatures'] = [float(x) for x in st.text_input("Temperatures (°C), Comma Separated", value="600, 650, 700, 750, 800, 850, 900, 950, 1000, 1050, 1075", key="ks_temperatures").split(",")]
        input_data['ks_values'] = [float(x) for x in st.text_input("Values (W/m.K), Comma Separated", value="36.5, 34.5, 32.6, 30.7, 29.0, 25.8, 26.2, 26.8, 27.4, 28.1, 28.4", key="ks_values").split(",")]

        st.write("Specific Heat (cs_data):")
        input_data['cs_temperatures'] = [float(x) for x in st.text_input("Temperatures (°C), Comma Separated", value="600, 650, 700, 750, 775, 800, 850, 900, 950, 1000, 1050, 1075", key="cs_temperatures").split(",")]
        input_data['cs_values'] = [float(x) for x in st.text_input( "Values (J/kg.K), Comma Separated", value="765, 824, 892, 971, 1014, 645, 647, 648, 650, 651, 653, 655", key="cs_values").split(",")]

        st.write("Density (ds_data):")
        input_data['ds_temperatures'] = [float(x) for x in st.text_input("Temperatures (°C), Comma Separated", value="600, 1300", key="ds_temperatures").split(",")]
        input_data['ds_values'] = [float(x) for x in st.text_input( "Values (kg/m3), Comma Separated", value="7800, 7332", key="ds_values").split(",")]

        st.write("Global Heat Transfer Coeffincient (hg_data):")
        input_data['hg_temperatures'] = [float(x) for x in st.text_input("Temperatures (°C), Comma Separated", value="700, 750, 800, 850, 900, 950, 1000, 1050, 1100, 1150, 1200", key="hg_temperatures").split(",")]
        input_data['hg_values'] = [float(x) for x in st.text_input( "Values (W/m^2.K), Comma Separated", value="64, 72, 80, 89, 98, 108, 120, 132, 145, 158, 173", key="hg_values").split(",")]

        st.write("Temperatures for calculating the cooling rate (°C):")
        start_temp = st.number_input("Initial temperature for calculating the cooling rate (°C)", value=820.0)
        end_temp = st.number_input("Final temperature for calculating the cooling rate (°C)", value=800.0)
    
    else:  # Upload de Arquivo XLSX
        st.subheader("Input Data (Upload XLSX File)")
        uploaded_file = st.file_uploader("Load XLSX File", type=["xlsx"])

        if uploaded_file is not None:
            try:
                # Ler o arquivo XLSX
                workbook = load_workbook(uploaded_file, data_only=True)
                sheet = workbook.active

                ks_temperatures = []
                ks_values = []
                cs_temperatures = []
                cs_values = []
                ds_temperatures = []
                ds_values = []
                hg_temperatures = []
                hg_values = []          

                # Extrair dados do arquivo XLSX
                input_data['len_x'] = float(sheet['D2'].value)
                input_data['len_y'] = float(sheet['D3'].value)
                input_data['len_w'] = float(sheet['D4'].value)
                input_data['len_z'] = float(sheet['D5'].value)
                input_data['min_temp'] = float(sheet['D6'].value)
                input_data['tamb'] = float(sheet['D7'].value)
                input_data['t_initial'] = float(sheet['D8'].value)

                number_ks = sheet['C9'].value

                for i in range(number_ks):
                    ks_temp = sheet.cell(row=9, column=4+i).value
                    ks_value = sheet.cell(row=10, column=4+i).value
                    ks_temperatures.append(ks_temp)
                    ks_values.append(ks_value)

                input_data['ks_temperatures'] = ks_temperatures
                input_data['ks_values'] = ks_values

                number_cs = sheet['C11'].value

                for i in range(number_cs):
                    cs_temp = sheet.cell(row=11, column=4+i).value
                    cs_value = sheet.cell(row=12, column=4+i).value
                    cs_temperatures.append(cs_temp)
                    cs_values.append(cs_value)

                input_data['cs_temperatures'] = cs_temperatures
                input_data['cs_values'] = cs_values

                number_ds = sheet['C13'].value

                for i in range(number_ds):
                    ds_temp = sheet.cell(row=13, column=4+i).value
                    ds_value = sheet.cell(row=14, column=4+i).value
                    ds_temperatures.append(ds_temp)
                    ds_values.append(ds_value)

                input_data['ds_temperatures'] = ds_temperatures
                input_data['ds_values'] = ds_values

                number_hg = sheet['C15'].value

                for i in range(number_hg):
                    hg_temp = sheet.cell(row=15, column=4+i).value
                    hg_value = sheet.cell(row=16, column=4+i).value
                    hg_temperatures.append(hg_temp)
                    hg_values.append(hg_value)

                input_data['hg_temperatures'] = hg_temperatures
                input_data['hg_values'] = hg_values

                start_temp = float(sheet['D17'].value)
                end_temp = float(sheet['D18'].value)

                st.success("XLSX file data loaded successfully!")

            except Exception as e:
                st.error(f"Error loading XLSX file: {e}")
                return  # Abortar se houver erro no upload

    # Botão de Calcular
    if st.button("Calculate Angle"):
        try:
            # Preparar os dados para a função de cálculo
            min_temp = input_data['min_temp']
            len_x = input_data['len_x']
            len_y = input_data['len_y']
            len_w = input_data['len_w']
            len_z = input_data['len_z']
            #delta = input_data['delta']
            delta = min(len_w/8, len_z/8)
            tamb = input_data['tamb']
            t_initial = input_data['t_initial']

            ks_data = (input_data['ks_temperatures'], input_data['ks_values'])
            cs_data = (input_data['cs_temperatures'], input_data['cs_values'])
            ds_data = (input_data['ds_temperatures'], input_data['ds_values'])
            hg_data = (input_data['hg_temperatures'], input_data['hg_values'])

            precalc_temps, precalc_ks, precalc_cs, precalc_ds, precalc_hg = precalculate_interpolated_properties(
                ks_data, cs_data, ds_data, hg_data
            )

            # Executar a simulação
            temperature_fields, temperaturas, dts, pontos, nx, ny, nz, nw = angle_heat_transfer_calculator(
                min_temp, len_x, len_y, len_w, len_z, delta,
                precalc_temps, precalc_ks, precalc_cs, precalc_ds, precalc_hg,  # Dados pré-calculados
                ks_data, cs_data, ds_data, hg_data,  # Dados originais (mantidos)
                tamb, t_initial
            )

            # Intervalo de tempo para salvar a animação (em segundos)
            animation_interval = 30

            # Chamar a função que cria a animação e calcula os tempos
            angle_tempos, angle_animation_path = create_animation_and_temps(temperature_fields, dts, len_x, len_y, delta, min_temp, t_initial, pontos, nx, ny, nz, animation_interval, "angle")

            # Exibir a animação no Streamlit
            st.subheader("Temperature Distribution Animation")
            try:
                st.image(angle_animation_path, use_container_width=True)
            except FileNotFoundError:
                st.error("Animation file not found. Please check if the animation was saved correctly.")
            except Exception as e:
                st.error(f"Error displaying animation: {e}")

            # Plotar a evolução da temperatura no tempo para os pontos monitorados
            st.subheader("Temperature Evolution at Monitored Points")
            fig_temp, ax_temp = plt.subplots()  # Criar uma nova figura para o gráfico de temperatura

            # Novos rótulos para os pontos
            # Access nx, ny, nz from the function's return
            point_labels = {
                (0, int(ny/2)): 'Leg 1 Surface',
                (int(nz / 2), int(2*ny / 3)): 'Leg 1 Middle',
                (int(2*nx/3), int(nw/2)): 'Leg 2 Middle'
            }

            for p in temperaturas:
                ax_temp.plot(angle_tempos, temperaturas[p], label=point_labels[p])

            ax_temp.set_xlabel('Time (seconds)')
            ax_temp.set_ylabel('Temperature (°C)')
            ax_temp.set_title('Temperature Evolution at Monitored Points')
            ax_temp.legend()
            st.pyplot(fig_temp)  # Exibir o gráfico no Streamlit

            # Armazenar os resultados no session state
            st.session_state['angle_tempos'] = angle_tempos
            st.session_state['angle_temperaturas'] = temperaturas
            st.session_state['angle_pontos'] = pontos
            st.session_state['angle_start_temp'] = start_temp
            st.session_state['angle_end_temp'] = end_temp
            st.session_state['angle_nx'] = nx  # Store nx in session state
            st.session_state['angle_ny'] = ny  # Store ny in session state
            st.session_state['angle_nz'] = nz  # Store nz in session state
            st.session_state['angle_nw'] = nw  # Store nz in session state

        except Exception as e:
            st.error(f"Error during calculation: {e}")

    # Opção para salvar os dados do gráfico em um arquivo TXT
    st.subheader("Save Temperature Evolution Data")
    if 'angle_tempos' in st.session_state and 'angle_temperaturas' in st.session_state and 'angle_pontos' in st.session_state:
        if st.button("Generate TXT File Angle"):
            angle_tempos = st.session_state['angle_tempos']
            angle_temperaturas = st.session_state['angle_temperaturas']
            angle_pontos = st.session_state['angle_pontos']
            t_initial = input_data['t_initial']
            min_temp = input_data['min_temp']
            temp_interval = (t_initial - min_temp)/12  # Intervalo de amostragem pegar 12 pontos entre o inicial e o final
            target_point_index = 1  # Índice do ponto de interesse (0, 1 ou 2)
            txt_data = save_temperature_data(angle_tempos, angle_temperaturas, angle_pontos, temp_interval, target_point_index)
            st.download_button(
                label="Download TXT file",
                data=txt_data,
                file_name="temperature_evolution_angle.txt",
                mime="text/plain",
            )
    else:
        st.write("Please run the calculation first to generate the TXT file.")

    # Cálculo da taxa de resfriamento
    st.subheader("Cooling Rate Calculation")
    if 'angle_tempos' in st.session_state and 'angle_temperaturas' in st.session_state and 'angle_pontos' in st.session_state and 'angle_nx' in st.session_state and 'angle_ny' in st.session_state and 'angle_nz' in st.session_state and 'angle_nw' in st.session_state:
        angle_tempos = st.session_state['angle_tempos']
        angle_temperaturas = st.session_state['angle_temperaturas']
        angle_pontos = st.session_state['angle_pontos']
        angle_start_temp = st.session_state['angle_start_temp']
        angle_end_temp = st.session_state['angle_end_temp']
        angle_nx = st.session_state['angle_nx']  # Retrieve nx from session state
        angle_ny = st.session_state['angle_ny']  # Retrieve ny from session state
        angle_nz = st.session_state['angle_nz']  # Retrieve nz from session state
        angle_nw = st.session_state['angle_nw']  # Retrieve nw from session state

        # Rótulos para os pontos
        point_labels_cooling = {
            (0, int(angle_ny/2)): 'Leg 1 Surface',
            (int(angle_nz/2), int(2*angle_ny/3)): 'Leg 1 Middle',
            (int(2*angle_nx/3), int(angle_nw/2)): 'Leg 2 Middle'
        }

        cooling_rates = {}
        for p in angle_pontos:
            cooling_rate = calculate_cooling_rate(angle_temperaturas[p], angle_tempos, angle_start_temp, angle_end_temp)
            cooling_rates[p] = cooling_rate

        st.write("Cooling rate (°C/s):")
        for p, rate in cooling_rates.items():
            st.write(f"{point_labels_cooling[p]}: {rate:.2f} °C/s")
    else:
        st.write("Please run the calculation first to see the cooling rates.")

def h_beam_calculator():
    # Limpar as variáveis de sessão específicas do Angle
    for key in ['angle_tempos', 'angle_temperaturas', 'angle_pontos', 'angle_start_temp', 'angle_end_temp', 'angle_nx', 'angle_ny', 'angle_nz', 'angle_nw']:
        if key in st.session_state:
            del st.session_state[key]

    st.title("H-Beam Heat Transfer Calculator")

    st.image("h-beam_dimensions_complete.png", use_container_width=True)

    # Opção de entrada de dados
    data_option = st.radio("Select data input type:", ("Typing", "Upload XLSX File"))

    # Inicializar dicionários para os dados
    input_data = {}

    if data_option == "Typing":
        st.subheader("Input Data (Typing)")
        # Dados de entrada numéricos
        input_data['len_x'] = st.number_input("Half Profile Height  (mm)", value=161.5, key="len_x")
        input_data['len_y'] = st.number_input("Half Flange Width (mm)", value=155.0, key="len_y")
        input_data['len_w'] = st.number_input("Half Web Thickness (mm)", value=7.0, key="len_w")
        input_data['len_z'] = st.number_input("Flange Thickness (mm)", value=22.9, key="len_z")
        #input_data['delta'] = st.number_input("Mesh Size (mm)", value=5.0, key="delta")
        input_data['min_temp'] = st.number_input("Final Temperature in Flange Middle (°C)", value=750.0, key="min_temp")
        input_data['tamb'] = st.number_input("Environment Temperature (°C)", value=30.0, key="tamb")
        input_data['t_initial'] = st.number_input("Initial Temperature (Finishing Rolling Temperature) (°C)", value=850.0, key="t_initial")

        # Dados de propriedades variáveis (listas de números)
        st.write("Thermal Conductivity (ks_data):")
        input_data['ks_temperatures'] = [float(x) for x in st.text_input("Temperatures (°C), Comma Separated", value="600, 650, 700, 750, 800, 850, 900, 950, 1000, 1050, 1075", key="ks_temperatures").split(",")]
        input_data['ks_values'] = [float(x) for x in st.text_input("Values (W/m.K), Comma Separated", value="36.5, 34.5, 32.6, 30.7, 29.0, 25.8, 26.2, 26.8, 27.4, 28.1, 28.4", key="ks_values").split(",")]

        st.write("Specific Heat (cs_data):")
        input_data['cs_temperatures'] = [float(x) for x in st.text_input("Temperatures (°C), Comma Separated", value="600, 650, 700, 750, 775, 800, 850, 900, 950, 1000, 1050, 1075", key="cs_temperatures").split(",")]
        input_data['cs_values'] = [float(x) for x in st.text_input( "Values (J/kg.K), Comma Separated", value="765, 824, 892, 971, 1014, 645, 647, 648, 650, 651, 653, 655", key="cs_values").split(",")]

        st.write("Density (ds_data):")
        input_data['ds_temperatures'] = [float(x) for x in st.text_input("Temperatures (°C), Comma Separated", value="600, 1300", key="ds_temperatures").split(",")]
        input_data['ds_values'] = [float(x) for x in st.text_input( "Values (kg/m3), Comma Separated", value="7800, 7332", key="ds_values").split(",")]

        st.write("Global Heat Transfer Coeffincient (hg_data):")
        input_data['hg_temperatures'] = [float(x) for x in st.text_input("Temperatures (°C), Comma Separated", value="700, 750, 800, 850, 900, 950, 1000, 1050, 1100, 1150, 1200", key="hg_temperatures").split(",")]
        input_data['hg_values'] = [float(x) for x in st.text_input( "Values (W/m^2.K), Comma Separated", value="64, 72, 80, 89, 98, 108, 120, 132, 145, 158, 173", key="hg_values").split(",")]

        st.write("Temperatures for calculating the cooling rate (°C):")
        start_temp = st.number_input("Initial temperature for calculating the cooling rate (°C)", value=820.0)
        end_temp = st.number_input("Final temperature for calculating the cooling rate (°C)", value=770.0)
    
    else:  # Upload de Arquivo XLSX
        st.subheader("Input Data (Upload XLSX File)")
        uploaded_file = st.file_uploader("Load XLSX File", type=["xlsx"])

        if uploaded_file is not None:
            try:
                # Ler o arquivo XLSX
                workbook = load_workbook(uploaded_file, data_only=True)
                sheet = workbook.active

                ks_temperatures = []
                ks_values = []
                cs_temperatures = []
                cs_values = []
                ds_temperatures = []
                ds_values = []
                hg_temperatures = []
                hg_values = []          

                # Extrair dados do arquivo XLSX
                input_data['len_x'] = float(sheet['D2'].value)
                input_data['len_y'] = float(sheet['D3'].value)
                input_data['len_w'] = float(sheet['D4'].value)
                input_data['len_z'] = float(sheet['D5'].value)
                input_data['min_temp'] = float(sheet['D6'].value)
                input_data['tamb'] = float(sheet['D7'].value)
                input_data['t_initial'] = float(sheet['D8'].value)

                number_ks = sheet['C9'].value

                for i in range(number_ks):
                    ks_temp = sheet.cell(row=9, column=4+i).value
                    ks_value = sheet.cell(row=10, column=4+i).value
                    ks_temperatures.append(ks_temp)
                    ks_values.append(ks_value)

                input_data['ks_temperatures'] = ks_temperatures
                input_data['ks_values'] = ks_values

                number_cs = sheet['C11'].value

                for i in range(number_cs):
                    cs_temp = sheet.cell(row=11, column=4+i).value
                    cs_value = sheet.cell(row=12, column=4+i).value
                    cs_temperatures.append(cs_temp)
                    cs_values.append(cs_value)

                input_data['cs_temperatures'] = cs_temperatures
                input_data['cs_values'] = cs_values

                number_ds = sheet['C13'].value

                for i in range(number_ds):
                    ds_temp = sheet.cell(row=13, column=4+i).value
                    ds_value = sheet.cell(row=14, column=4+i).value
                    ds_temperatures.append(ds_temp)
                    ds_values.append(ds_value)

                input_data['ds_temperatures'] = ds_temperatures
                input_data['ds_values'] = ds_values

                number_hg = sheet['C15'].value

                for i in range(number_hg):
                    hg_temp = sheet.cell(row=15, column=4+i).value
                    hg_value = sheet.cell(row=16, column=4+i).value
                    hg_temperatures.append(hg_temp)
                    hg_values.append(hg_value)

                input_data['hg_temperatures'] = hg_temperatures
                input_data['hg_values'] = hg_values

                start_temp = float(sheet['D17'].value)
                end_temp = float(sheet['D18'].value)

                st.success("XLSX file data loaded successfully!")

            except Exception as e:
                st.error(f"Error loading XLSX file: {e}")
                return  # Abortar se houver erro no upload

    # Botão de Calcular
    if st.button("Calculate H-Beam"):
        try:
            # Preparar os dados para a função de cálculo
            min_temp = input_data['min_temp']
            len_x = input_data['len_x']
            len_y = input_data['len_y']
            len_w = input_data['len_w']
            len_z = input_data['len_z']
            #delta = input_data['delta']
            delta = min(len_w/4, len_z/4)
            tamb = input_data['tamb']
            t_initial = input_data['t_initial']

            ks_data = (input_data['ks_temperatures'], input_data['ks_values'])
            cs_data = (input_data['cs_temperatures'], input_data['cs_values'])
            ds_data = (input_data['ds_temperatures'], input_data['ds_values'])
            hg_data = (input_data['hg_temperatures'], input_data['hg_values'])

            precalc_temps, precalc_ks, precalc_cs, precalc_ds, precalc_hg = precalculate_interpolated_properties(
                ks_data, cs_data, ds_data, hg_data
            )

            # Executar a simulação
            temperature_fields, temperaturas, dts, pontos, nx, ny, nz = h_beam_heat_transfer_calculator(
                min_temp, len_x, len_y, len_w, len_z, delta,
                precalc_temps, precalc_ks, precalc_cs, precalc_ds, precalc_hg,  # Dados pré-calculados
                ks_data, cs_data, ds_data, hg_data,  # Dados originais (mantidos)
                tamb, t_initial
            )

            # Intervalo de tempo para salvar a animação (em segundos)
            animation_interval = 30

            # Chamar a função que cria a animação e calcula os tempos
            hbeam_tempos, hbeam_animation_path = create_animation_and_temps(temperature_fields, dts, len_x, len_y, delta, min_temp, t_initial, pontos, nx, ny, nz, animation_interval, "h_beam")

            # Exibir a animação no Streamlit
            st.subheader("Temperature Distribution Animation")
            try:
                st.image(hbeam_animation_path, use_container_width=True)
            except FileNotFoundError:
                st.error("Animation file not found. Please check if the animation was saved correctly.")
            except Exception as e:
                st.error(f"Error displaying animation: {e}")

            # Plotar a evolução da temperatura no tempo para os pontos monitorados
            st.subheader("Temperature Evolution at Monitored Points")
            fig_temp, ax_temp = plt.subplots()  # Criar uma nova figura para o gráfico de temperatura

            # Novos rótulos para os pontos
            # Access nx, ny, nz from the function's return
            point_labels = {
                (0, 0): 'Flange Surface',
                (int(nz / 2), int(2 * ny / 3)): 'Flange Middle',
                (int(nx - 1), 0): 'Web Middle'
            }

            for p in temperaturas:
                ax_temp.plot(hbeam_tempos, temperaturas[p], label=point_labels[p])

            ax_temp.set_xlabel('Time (seconds)')
            ax_temp.set_ylabel('Temperature (°C)')
            ax_temp.set_title('Temperature Evolution at Monitored Points')
            ax_temp.legend()
            st.pyplot(fig_temp)  # Exibir o gráfico no Streamlit

            # Armazenar os resultados no session state
            st.session_state['hbeam_tempos'] = hbeam_tempos
            st.session_state['hbeam_temperaturas'] = temperaturas
            st.session_state['hbeam_pontos'] = pontos
            st.session_state['hbeam_start_temp'] = start_temp
            st.session_state['hbeam_end_temp'] = end_temp
            st.session_state['hbeam_nx'] = nx  # Store nx in session state
            st.session_state['hbeam_ny'] = ny  # Store ny in session state
            st.session_state['hbeam_nz'] = nz  # Store nz in session state
            #st.session_state['hbeam_nw'] = nw # Store nw in session state

        except Exception as e:
            st.error(f"Error during calculation: {e}")

    # Opção para salvar os dados do gráfico em um arquivo TXT
    st.subheader("Save Temperature Evolution Data")
    if 'hbeam_tempos' in st.session_state and 'hbeam_temperaturas' in st.session_state and 'hbeam_pontos' in st.session_state:
        if st.button("Generate TXT File H-Beam"):
            hbeam_tempos = st.session_state['hbeam_tempos']
            hbeam_temperaturas = st.session_state['hbeam_temperaturas']
            hbeam_pontos = st.session_state['hbeam_pontos']
            t_initial = input_data['t_initial']
            min_temp = input_data['min_temp']
            temp_interval = (t_initial - min_temp)/12  # Intervalo de amostragem pegar 12 pontos entre o inicial e o final
            target_point_index = 1  # Índice do ponto de interesse (0, 1 ou 2)
            txt_data = save_temperature_data(hbeam_tempos, hbeam_temperaturas, hbeam_pontos, temp_interval, target_point_index)
            st.download_button(
                label="Download TXT file",
                data=txt_data,
                file_name="temperature_evolution_hbeam.txt",
                mime="text/plain",
            )
    else:
        st.write("Please run the calculation first to generate the TXT file.")

    # Cálculo da taxa de resfriamento
    st.subheader("Cooling Rate Calculation")
    if 'hbeam_tempos' in st.session_state and 'hbeam_temperaturas' in st.session_state and 'hbeam_pontos' in st.session_state and 'hbeam_nx' in st.session_state and 'hbeam_ny' in st.session_state and 'hbeam_nz' in st.session_state:
        hbeam_tempos = st.session_state['hbeam_tempos']
        hbeam_temperaturas = st.session_state['hbeam_temperaturas']
        hbeam_pontos = st.session_state['hbeam_pontos']
        hbeam_start_temp = st.session_state['hbeam_start_temp']
        hbeam_end_temp = st.session_state['hbeam_end_temp']
        hbeam_nx = st.session_state['hbeam_nx']  # Retrieve nx from session state
        hbeam_ny = st.session_state['hbeam_ny']  # Retrieve ny from session state
        hbeam_nz = st.session_state['hbeam_nz']  # Retrieve nz from session state

        # Rótulos para os pontos
        point_labels_cooling = {
            (0, 0): 'Flange Surface',
            (int(hbeam_nz/2), int(2*hbeam_ny/3)): 'Flange Middle',
            (int(hbeam_nx-1), 0): 'Web Middle'
        }

        cooling_rates = {}
        for p in hbeam_pontos:
            cooling_rate = calculate_cooling_rate(hbeam_temperaturas[p], hbeam_tempos, hbeam_start_temp, hbeam_end_temp)
            cooling_rates[p] = cooling_rate

        st.write("Cooling rate (°C/s):")
        for p, rate in cooling_rates.items():
            st.write(f"{point_labels_cooling[p]}: {rate:.2f} °C/s")
    else:
        st.write("Please run the calculation first to see the cooling rates.")

# Sidebar para navegação
menu = ["Login", "Angle Heat Transfer Calculator", "H-Beam Heat Transfer Calculator"]
choice = st.sidebar.selectbox("Menu", menu)

# Lógica de roteamento
if st.session_state['authentication_status'] != True:
    login()
else:
    if choice == "Angle Heat Transfer Calculator":
        angle_calculator()
    elif choice == "H-Beam Heat Transfer Calculator":
        h_beam_calculator()
    elif choice == "Login":
        st.write(f"Welcome, {st.session_state['username']}!")
        if st.button("Logout"):
            st.session_state['authentication_status'] = False
            st.session_state['username'] = None